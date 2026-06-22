import json
import os
import re
import shutil
from io import BytesIO
from typing import Literal

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, WebSocket, WebSocketDisconnect
from pydantic import BaseModel
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app import crud
from app.database import get_db
from app.services.pipeline_service import process_student_card

router = APIRouter(prefix="/attendance", tags=["Attendance"])


# --- TRẠM QUẢN LÝ WEBSOCKET ---
class ConnectionManager:
    def __init__(self):
        self.active_connections: list[WebSocket] = []

    async def connect(self, websocket: WebSocket):
        await websocket.accept()
        self.active_connections.append(websocket)

    def disconnect(self, websocket: WebSocket):
        if websocket in self.active_connections:
            self.active_connections.remove(websocket)

    async def broadcast(self, message: str):
        for connection in self.active_connections:
            await connection.send_text(message)


manager = ConnectionManager()


def format_excel_datetime(value):
    if not value:
        return "-"

    return value.strftime("%d/%m/%Y %H:%M:%S")


def safe_excel_filename(value: str, fallback: str):
    normalized = re.sub(r"[^a-zA-Z0-9_-]+", "-", value.strip()).strip("-")
    return normalized.lower() or fallback


# Mở đường ống WebSocket tại địa chỉ: ws://127.0.0.1:8000/attendance/ws
@router.websocket("/ws")
async def websocket_endpoint(websocket: WebSocket):
    await manager.connect(websocket)
    try:
        while True:
            await websocket.receive_text()
    except WebSocketDisconnect:
        manager.disconnect(websocket)


@router.get("/active-session")
def read_active_session(db: Session = Depends(get_db)):
    active_session = crud.get_active_session(db)
    if not active_session:
        return None

    return crud.get_session_summary(db, active_session)


@router.get("/sessions")
def read_attendance_sessions(db: Session = Depends(get_db)):
    return crud.get_all_session_summaries(db)


class AttendanceSessionStartRequest(BaseModel):
    name: str | None = None


@router.post("/sessions")
async def start_attendance_session(
    request: AttendanceSessionStartRequest | None = None,
    db: Session = Depends(get_db),
):
    try:
        session_name = request.name.strip() if request and request.name else None
        active_session = crud.start_attendance_session(db, name=session_name)
        session_summary = crud.get_session_summary(db, active_session)

        await manager.broadcast(json.dumps({
            "type": "SESSION_STARTED",
            "session_id": active_session.id,
        }))

        return {
            "success": True,
            "message": "Đã khởi tạo buổi học",
            "session": session_summary,
        }
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/history/{session_id}")
def read_attendance_history(session_id: int, db: Session = Depends(get_db)):
    attendance_session = crud.get_attendance_session(db, session_id=session_id)
    if not attendance_session:
        raise HTTPException(status_code=404, detail="Không tìm thấy buổi học")

    return {
        "session": crud.get_session_summary(db, attendance_session, active_only=False),
        "students": crud.get_students_status(db, session_id=session_id, active_only=False),
    }


@router.get("/history/{session_id}/export")
def export_attendance_history(session_id: int, db: Session = Depends(get_db)):
    attendance_session = crud.get_attendance_session(db, session_id=session_id)
    if not attendance_session:
        raise HTTPException(status_code=404, detail="Không tìm thấy buổi học")

    session_summary = crud.get_session_summary(db, attendance_session, active_only=False)
    students = crud.get_students_status(db, session_id=session_id, active_only=False)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Diem danh"

    title_fill = PatternFill("solid", fgColor="D12027")
    header_fill = PatternFill("solid", fgColor="FCE4E4")
    title_font = Font(color="FFFFFF", bold=True, size=14)
    header_font = Font(bold=True, color="2D3748")
    center_alignment = Alignment(horizontal="center", vertical="center")

    sheet.merge_cells("A1:E1")
    sheet["A1"] = "KET QUA DIEM DANH"
    sheet["A1"].font = title_font
    sheet["A1"].fill = title_fill
    sheet["A1"].alignment = center_alignment

    session_rows = [
        ("Buổi học", session_summary["name"]),
        ("Bắt đầu", format_excel_datetime(session_summary["started_at"])),
        ("Kết thúc", format_excel_datetime(session_summary["ended_at"])),
        ("Sĩ số", session_summary["total_students"]),
        ("Có mặt", session_summary["present_count"]),
        ("Đi muộn", session_summary["late_count"]),
        ("Vắng", session_summary["absent_count"]),
    ]

    for row_index, (label, value) in enumerate(session_rows, start=3):
        sheet.cell(row=row_index, column=1, value=label).font = Font(bold=True)
        sheet.cell(row=row_index, column=2, value=value)

    header_row = 11
    headers = ["STT", "Mã sinh viên", "Họ và tên", "Trạng thái", "Thời gian điểm danh"]
    for column_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=header_row, column=column_index, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment

    for index, student in enumerate(students, start=1):
        row_index = header_row + index
        is_present = student["is_present"]
        status_label = student.get("status_label", "Có mặt" if is_present else "Vắng")
        row_values = [
            index,
            student["student_code"],
            student["name"],
            status_label,
            format_excel_datetime(student["checked_in_at"]) if is_present else "-",
        ]

        for column_index, value in enumerate(row_values, start=1):
            cell = sheet.cell(row=row_index, column=column_index, value=value)
            if column_index in (1, 4, 5):
                cell.alignment = center_alignment

    column_widths = [8, 18, 28, 16, 24]
    for column_index, width in enumerate(column_widths, start=1):
        sheet.column_dimensions[get_column_letter(column_index)].width = width

    sheet.freeze_panes = f"A{header_row + 1}"
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    filename = f"diem-danh-{safe_excel_filename(session_summary['name'], str(session_id))}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# --- CÁC API XỬ LÝ ĐIỂM DANH ---
@router.post("/checkin")
async def checkin(file: UploadFile = File(...), db: Session = Depends(get_db)):
    temp_file_path = f"temp_{file.filename}"
    try:
        with open(temp_file_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)

        active_session = crud.get_active_session(db)
        if not active_session:
            return {
                "success": False,
                "code": "NO_ACTIVE_SESSION",
                "message": "Chưa khởi tạo buổi học",
            }

        student_code = process_student_card(temp_file_path)

        if not student_code or student_code == "Không nhận diện được":
            return {"success": False, "message": "Không đọc được mã sinh viên, vui lòng thử lại!"}

        student = crud.get_student_by_code(db, student_code=student_code)
        if not student:
            return {
                "success": False,
                "code": "STUDENT_NOT_IN_CLASS",
                "student_code": student_code,
                "message": "Sinh viên không có trong lớp",
            }

        if crud.has_attended(db, student_code=student_code, session_id=active_session.id):
            return {
                "success": False,
                "code": "ALREADY_CHECKED_IN",
                "student_code": student.student_code,
                "session_id": active_session.id,
                "message": "Bạn đã điểm danh rồi",
            }

        crud.record_attendance(
            db,
            student_code=student_code,
            session_id=active_session.id,
            status=crud.ATTENDANCE_STATUS_PRESENT,
        )

        await manager.broadcast(json.dumps({
            "type": "CHECKIN_SUCCESS",
            "student_code": student.student_code,
            "session_id": active_session.id,
            "status": crud.ATTENDANCE_STATUS_PRESENT,
        }))

        return {
            "success": True,
            "message": "Điểm danh thành công",
            "student_code": student.student_code,
            "name": student.name,
            "status": crud.ATTENDANCE_STATUS_PRESENT,
            "status_label": crud.ATTENDANCE_STATUS_LABELS[crud.ATTENDANCE_STATUS_PRESENT],
            "session_id": active_session.id,
            "session_name": active_session.name,
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

    finally:
        if os.path.exists(temp_file_path):
            os.remove(temp_file_path)


class ManualCheckinRequest(BaseModel):
    student_code: str
    status: Literal["present", "late"] = "present"


@router.post("/manual")
async def manual_checkin(request: ManualCheckinRequest, db: Session = Depends(get_db)):
    try:
        student_code = request.student_code.strip().upper()
        attendance_status = crud.normalize_attendance_status(request.status)
        student = crud.get_student_by_code(db, student_code=student_code)
        if not student:
            raise HTTPException(status_code=404, detail="Không tìm thấy sinh viên")

        active_session = crud.get_active_session(db)
        if not active_session:
            raise HTTPException(status_code=400, detail="Chưa khởi tạo buổi học")

        if crud.has_attended(db, student_code=student_code, session_id=active_session.id):
            raise HTTPException(status_code=400, detail="Bạn đã điểm danh rồi")

        crud.record_attendance(
            db,
            student_code=student_code,
            session_id=active_session.id,
            status=attendance_status,
        )

        await manager.broadcast(json.dumps({
            "type": "CHECKIN_SUCCESS",
            "student_code": student_code,
            "session_id": active_session.id,
            "status": attendance_status,
        }))

        status_label = crud.ATTENDANCE_STATUS_LABELS[attendance_status]
        return {
            "success": True,
            "message": f"Đã ghi nhận {status_label.lower()} cho {student_code}",
            "status": attendance_status,
            "status_label": status_label,
            "session_id": active_session.id,
            "session_name": active_session.name,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/reset")
async def reset_all_attendance(db: Session = Depends(get_db)):
    try:
        closed_session = crud.reset_attendance(db)
        if not closed_session:
            raise HTTPException(status_code=400, detail="Không có buổi học đang mở")

        await manager.broadcast(json.dumps({
            "type": "RESET_SUCCESS",
            "closed_session_id": closed_session.id,
        }))

        return {
            "success": True,
            "message": "Đã kết thúc buổi học và lưu lịch sử điểm danh",
            "closed_session": crud.get_session_summary(db, closed_session),
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
